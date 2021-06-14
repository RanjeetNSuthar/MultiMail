from email.encoders import encode_base64
import re
import openpyxl
import auto_attachment

import smtplib
import os
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from gui_multimail import Ui_Form
from PyQt5.QtWidgets import QDialog,QMessageBox,QPushButton


class Mailer(QDialog,Ui_Form):
    def __init__(self):
        super(Mailer, self).__init__()
        self.setupUi(self)
        self.send_button.clicked.connect(self.on_send)
        self.browse_button_excel.clicked.connect(self.on_browse_excel)
        self.browse_button_mail.clicked.connect(self.on_browse_mail)
        self.Open_word_button.clicked.connect(self.on_create_file)
        self.Open_excel_button.clicked.connect(self.on_resource)
    
    def on_send(self):


        '''
        receipt_no = str(self.receipt_no_text.text())
        name = str(self.name_text.text())
        junior_senior = str(self.junior_senior_text.text())
        college = str(self.college_text.text())
        contact_no = str(self.contact_no_text.text())
        email_address = str(self.email_address_text.text())
        '''
        
        # Get filename - FILENAME
        # Get row number of emails in CSV file - row_num
        # Get email id and password - EMAIL_ID, PASSWORD
        # Get subject text - SUBJECT_TEXT
        # Get mail content - MAIL_CONTENT
        
        path_excel = "C://Users\Admi\Desktop\SEM-4-IT\MiniProject_sem4\git_project\miniproject/email&attachment_list.xlsx"
        column_number = 3
        EMAIL_ID = str(self.email_id_text.text())
        PASSWORD = str(self.password_text.text())
        SUBJECT_TEXT = str(self.subject_text.text())
        path_mail = "C://Users\Admi\Desktop\SEM-4-IT\MiniProject_sem4\git_project\miniproject/body.txt"
        MAIL_CONTENT = 'initialized_value'
        

        msg = MIMEMultipart()
        msg['From'] = EMAIL_ID
        msg['Subject'] = SUBJECT_TEXT
        with open (path_mail, "r") as myfile:
            MAIL_CONTENT = myfile.read()

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(EMAIL_ID, PASSWORD)

        wb = openpyxl.load_workbook(path_excel)
        sh5 = wb['Sheet1']
        for rows in range(2, sh5.max_row + 1):
            msg['Date'] = formatdate(localtime=True)
            msg['To'] = recipient = str(sh5.cell(rows, column_number).value)

            replacement =str(sh5.cell(rows, column_number-1).value)
            print(rows,column_number-1)
            text = self.modified_mc(MAIL_CONTENT,replacement)


            attachment_filename = str(sh5.cell(rows, (column_number + 1)).value)
            print(rows,column_number + 1)
            encoded_attach = self.attachment_encoder(attachment_filename)
            print("Encoded")
            msg.attach(encoded_attach)
            msg.attach(MIMEText(text))
            server.sendmail(EMAIL_ID, recipient, msg.as_string())
            print(text)
            #print(encoded_attach)
            msg.set_payload(None)


        server.quit()
        print('done')
        msgBox = QMessageBox()
        msgBox.setText('Finish')
        msgBox.addButton(QPushButton('Ok'), QMessageBox.YesRole)
        ret = msgBox.exec_();


        # TO SEND IMAGES
        #fp = open(PATH_TO_IMAGE, 'rb')
        #img = MIMEImage(fp.read(),name=os.path.basename(PATH_TO_IMAGE))
        #fp.close()
        #msg.attach(img)

        # PARTS = int(ceil(len(total_recipients)/100.0))
        # print(PARTS)
        # for i in range(PARTS):
        #     print('{0} parts remaining..'.format(PARTS-i))
        #     recipients = total_recipients[i*100: min(len(total_recipients), (i+1)*100)]

    def on_browse_excel(self):
        run_obj1 = auto_attachment.runing_files()
        run_obj1.create_personalized_attachments()
        print("After personalize")
        os.startfile("C://Users\Admi\Desktop\SEM-4-IT\MiniProject_sem4\git_project\miniproject/email&attachment_list.xlsx")
    
    def on_browse_mail(self):
        os.startfile("C:/Users\Admi\Desktop\SEM-4-IT\MiniProject_sem4\git_project\miniproject/body.txt")
    
    def on_create_file(self):
        os.startfile("C:/Users\Admi\Desktop\SEM-4-IT\MiniProject_sem4\git_project\miniproject/Template.docx",)
    
    def on_resource(self):
        print("inside resource")
        try:
            run_obj = auto_attachment.runing_files()
            run_obj.create_resource_file()
        except Exception as e5:
            print(e5)
        os.startfile("C:/Users\Admi\Desktop\SEM-4-IT\MiniProject_sem4\git_project\miniproject/Book1.xlsx")

    def modified_mc(self,mail_content,replacement):
        pattern = re.compile(r"([{])([a-zA-Z]*)([}])")
        text = pattern.sub(replacement,mail_content)
        return text


    def attachment_encoder(self, filename):
        print("inside attach")
        f = filename
        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(f, "rb").read())
        encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(f))
        return part

